from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import os
import shutil


class Export:

    @staticmethod
    def limpiar_directorio_result():
        ruta_result = os.path.join(os.getcwd(), "result")
        if os.path.exists(ruta_result):
            for archivo in os.listdir(ruta_result):
                ruta_completa = os.path.join(ruta_result, archivo)
                if os.path.isfile(ruta_completa) or os.path.islink(ruta_completa):
                    os.remove(ruta_completa)
                elif os.path.isdir(ruta_completa):
                    shutil.rmtree(ruta_completa)

    @staticmethod
    def matriculas(listaAlumnos, seleccion, cruce, sustitucion,carpeta, script, nsga3 = False, conf = None, i = 0):
        if (not script and nsga3 and i == 1) or (not script and not nsga3):
            Export.limpiar_directorio_result()
        wb = Workbook()
        ws = wb.active
        ws.append(["AÑO", "PLAN", "CODIGO", "ASIGNATURA", "ALUMNO", "GRUPO", "GP"])
        for alumno in listaAlumnos:
            matriculas = alumno.matriculas_variables + alumno.matriculas_fijas
            for matricula in matriculas:
                ws.append(["2023-24", "406", matricula.cod_asignatura, matricula.nombre_asignatura, alumno.nombre,
                           matricula.grupo, matricula.grupo_practicas])

        ws.column_dimensions[get_column_letter(4)].width = len(
            "PLANIFICACIÓN E INTEGRACIÓN DE SISTEMAS Y SERVICIOS") + 0.5
        ws.column_dimensions[get_column_letter(5)].width = len("Estudiante000") + 0.4
        if not nsga3:
            if script:
                wb.save(carpeta+'/solucion-'+seleccion+"-"+cruce+"-"+sustitucion+".xlsx")
            else:
                wb.save(carpeta + '/matriculas.xlsx')
        else:
            if script and False:
                wb.save(carpeta + '/solucion' + str(conf.solapes) + "-" + str(conf.tasa_cohesion) + "-" + str(conf.d_total) + "-" + str(conf.tasa_practicas_pronto) + "-" + str(conf.tasa_cohesion_practicas) + "-" + str(conf.tasa_preferencias) + ".xlsx")
            else:
                os.makedirs(carpeta+"/Solucion_" + str(i), exist_ok=True)
                wb.save(carpeta + "/Solucion_" + str(i) + '/matriculas.xlsx')
    @staticmethod
    def alumnos_clase(alumnosClaseInicio, alumnosClaseFinal, asignaturas, seleccion, cruce, sustitucion, carpeta, script, nsga3 = False, conf = None, i=0):

        wb = Workbook()

        hoja1 = wb.active
        hoja1.title = "Configuracion Inicial"
        hoja2 = wb.create_sheet(title="Configuracion Final")

        hoja1.append(
            ["NOMBRE", "ALUMNOS", "ESPERADOS POR GRUPO TEORIA", "ESPERADOS POR GRUPO PRACTICAS", "GRUPO 10", "", "GP1", "",
             "GP2", "", "GRUPO 11", "", "GP1", "", "GP2", "", "GRUPO 12", "", "GP1", "", "GP2", "", ])
        hoja2.append(
            ["NOMBRE", "ALUMNOS", "ESPERADOS POR GRUPO TEORIA", "ESPERADOS POR GRUPO PRACTICAS", "GRUPO 10", "", "GP1",
             "", "GP2", "", "GRUPO 11", "", "GP1", "", "GP2", "", "GRUPO 12", "", "GP1", "", "GP2", "", ])

        alumnosClaseTeoriaInicio, alumnosClasePracticasInicio = alumnosClaseInicio[0], alumnosClaseInicio[1]
        alumnosClaseTeoriaFinal, alumnosClasePracticasFinal = alumnosClaseFinal[0], alumnosClaseFinal[1]

        percentage_style = NamedStyle(name="percentage_style", number_format="0.00%")

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        for clave in alumnosClaseTeoriaInicio.keys():

            cod_asignatura = clave[0]
            filtro = (asignaturas["COD. ASIG"] == cod_asignatura)
            datos_asignatura = asignaturas.loc[filtro, ["NOMBRE ASIGNATURA", "CURSO", "COD. ASIG"]]
            alumnos_teoria_10 = alumnosClaseTeoriaInicio[clave][10]
            alumnos_teoria_10_1 = alumnosClasePracticasInicio[clave[0]][10][1]
            alumnos_teoria_10_2 = alumnosClasePracticasInicio[clave[0]][10][2]
            alumnos_teoria_11 = alumnosClaseTeoriaInicio[clave][11]
            alumnos_teoria_11_1 = alumnosClasePracticasInicio[clave[0]][11][1]
            alumnos_teoria_11_2 = alumnosClasePracticasInicio[clave[0]][11][2]
            alumnos_teoria_12 = alumnosClaseTeoriaInicio[clave][12]
            alumnos_teoria_12_1 = alumnosClasePracticasInicio[clave[0]][12][1]
            alumnos_teoria_12_2 = alumnosClasePracticasInicio[clave[0]][12][2]

            alumnos_teoria_10_f = alumnosClaseTeoriaFinal[clave][10]
            alumnos_teoria_10_1_f = alumnosClasePracticasFinal[clave[0]][10][1]
            alumnos_teoria_10_2_f = alumnosClasePracticasFinal[clave[0]][10][2]
            alumnos_teoria_11_f = alumnosClaseTeoriaFinal[clave][11]
            alumnos_teoria_11_1_f = alumnosClasePracticasFinal[clave[0]][11][1]
            alumnos_teoria_11_2_f = alumnosClasePracticasFinal[clave[0]][11][2]
            alumnos_teoria_12_f = alumnosClaseTeoriaFinal[clave][12]
            alumnos_teoria_12_1_f = alumnosClasePracticasFinal[clave[0]][12][1]
            alumnos_teoria_12_2_f = alumnosClasePracticasFinal[clave[0]][12][2]

            if ((datos_asignatura["CURSO"] == 1) | (datos_asignatura["COD. ASIG"] == 42325)).any():
                alumnos_teoria = int(sum(alumnosClaseTeoriaInicio[clave].values()) / 3)
                alumnos_practica = int((sum(alumnosClaseTeoriaInicio[clave].values()) / 3) / 2)
                hoja1.append(
                    [datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaInicio[clave].values()),
                     alumnos_teoria,
                     alumnos_practica, alumnos_teoria_10, abs(alumnos_teoria_10 - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_10_1, abs(alumnos_teoria_10_1 - alumnos_practica) / alumnos_practica, alumnos_teoria_10_2, abs(alumnos_teoria_10_2 - alumnos_practica) / alumnos_practica,
                     alumnos_teoria_11, abs(alumnos_teoria_11 - alumnos_teoria) / alumnos_teoria, alumnos_teoria_11_1, abs(alumnos_teoria_11 - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_11_2, abs(alumnos_teoria_11_1 - alumnos_practica) / alumnos_practica, alumnos_teoria_12, abs(alumnos_teoria_12 - alumnos_teoria) / alumnos_teoria, alumnos_teoria_12_1,
                     abs(alumnos_teoria_12_1 - alumnos_practica) / alumnos_practica, alumnos_teoria_12_2, abs(alumnos_teoria_12_2 - alumnos_practica) / alumnos_practica])
                hoja2.append(
                    [datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaInicio[clave].values()),
                     alumnos_teoria,
                     alumnos_practica, alumnos_teoria_10_f, abs(alumnos_teoria_10_f - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_10_1_f, abs(alumnos_teoria_10_1_f - alumnos_practica) / alumnos_practica, alumnos_teoria_10_2_f, abs(alumnos_teoria_10_2_f - alumnos_practica) / alumnos_practica,
                     alumnos_teoria_11_f, abs(alumnos_teoria_11_f - alumnos_teoria) / alumnos_teoria, alumnos_teoria_11_1_f, abs(alumnos_teoria_11_f - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_11_2_f, abs(alumnos_teoria_11_2_f - alumnos_practica) / alumnos_practica, alumnos_teoria_12_f, abs(alumnos_teoria_12_f - alumnos_teoria) / alumnos_teoria, alumnos_teoria_12_1_f,
                     abs(alumnos_teoria_12_1_f - alumnos_practica) / alumnos_practica, alumnos_teoria_12_2_f, abs(alumnos_teoria_12_2_f - alumnos_practica) / alumnos_practica])
            else:
                alumnos_teoria = int(sum(alumnosClaseTeoriaInicio[clave].values()) / 2)
                alumnos_practica = int((sum(alumnosClaseTeoriaInicio[clave].values()) / 2) / 2)
                hoja1.append(
                    [datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaInicio[clave].values()),
                     alumnos_teoria,
                     alumnos_practica, alumnos_teoria_10, abs(alumnos_teoria_10 - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_10_1, abs(alumnos_teoria_10_1 - alumnos_practica) / alumnos_practica,
                     alumnos_teoria_10_2, abs(alumnos_teoria_10_2 - alumnos_practica) / alumnos_practica,
                     alumnos_teoria_11, abs(alumnos_teoria_11 - alumnos_teoria) / alumnos_teoria, alumnos_teoria_11_1,
                     abs(alumnos_teoria_11 - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_11_2, abs(alumnos_teoria_11_1 - alumnos_practica) / alumnos_practica,
                     "", "", "",
                     "", "",
                     ""])
                hoja2.append(
                    [datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaInicio[clave].values()),
                     alumnos_teoria,
                     alumnos_practica, alumnos_teoria_10_f, abs(alumnos_teoria_10_f - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_10_1_f, abs(alumnos_teoria_10_1_f - alumnos_practica) / alumnos_practica,
                     alumnos_teoria_10_2_f, abs(alumnos_teoria_10_2_f - alumnos_practica) / alumnos_practica,
                     alumnos_teoria_11_f, abs(alumnos_teoria_11_f - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_11_1_f, abs(alumnos_teoria_11_f - alumnos_teoria) / alumnos_teoria,
                     alumnos_teoria_11_2_f, abs(alumnos_teoria_11_2_f - alumnos_practica) / alumnos_practica,
                     "", "", "",
                     "", "",
                     ""])
            for col_idx in [6, 8, 10, 12, 14, 16, 18, 20, 22]:
                hoja1.cell(row=hoja1.max_row, column=col_idx).style = percentage_style
                hoja2.cell(row=hoja2.max_row, column=col_idx).style = percentage_style

            for hoja in [hoja1, hoja2]:
                for col_idx in [6, 8, 10, 12, 14, 16, 18, 20, 22]:
                    cell_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{hoja.max_row}"
                    rule = CellIsRule(operator="greaterThan", formula=["0.15"], stopIfTrue=True, fill=red_fill,
                                      font=red_font)
                    hoja.conditional_formatting.add(cell_range, rule)


        hoja1.column_dimensions[get_column_letter(1)].width = len("PROGRAMACIÓN CONCURRENTE Y TIEMPO REAL") + 5.7
        hoja1.column_dimensions[get_column_letter(2)].width = len("ALUMNOS") + 3
        hoja1.column_dimensions[get_column_letter(3)].width = len("ESPERADOS POR GRUPO TEORIA") + 3.3
        hoja1.column_dimensions[get_column_letter(4)].width = len("ESPERADOS POR GRUPO PRACTICAS") + 3.3

        hoja2.column_dimensions[get_column_letter(1)].width = len("PROGRAMACIÓN CONCURRENTE Y TIEMPO REAL") + 5.7
        hoja2.column_dimensions[get_column_letter(2)].width = len("ALUMNOS") + 3
        hoja2.column_dimensions[get_column_letter(3)].width = len("ESPERADOS POR GRUPO TEORIA") + 3.3
        hoja2.column_dimensions[get_column_letter(4)].width = len("ESPERADOS POR GRUPO PRACTICAS") + 3.3

        if not nsga3:
            if script:
                wb.save(carpeta+"/"+seleccion+"-"+cruce+"-"+sustitucion+".xlsx")
            else:
                wb.save(carpeta + "/distribucion-grupos.xlsx")
        else:
            if script and False:
                wb.save(carpeta + "/alumnos" + str(conf.solapes) + "-" + str(conf.tasa_cohesion) + "-" + str(conf.d_total) + "-" + str(conf.tasa_practicas_pronto) + "-" + str(conf.tasa_cohesion_practicas) + "-" + str(conf.tasa_preferencias) + ".xlsx")
            else:
                os.makedirs(carpeta + "/Solucion_" + str(i), exist_ok=True)
                wb.save(carpeta +  "/Solucion_" + str(i) + "/distribucion-grupos.xlsx")
