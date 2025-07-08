import time

import multiprocessing as mp
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

from utils.utils import Utils

from model.nsga3 import NSGA3
from utils.utils import Utils

# Este script realiza procesamiento intensivo y genera múltiples archivos de gran tamaño.
# ⚠️ Tiempo estimado de ejecución: más de 9 horas.
# Requiere una cantidad significativa de espacio en disco para almacenar los resultados.
# Asegúrate de contar con suficiente almacenamiento disponible antes de iniciar la ejecución.

def apply_border(ws, start_row, start_col, end_row, end_col, border):
    """Aplica bordes a un rango de celdas."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border


def create_excel_nsg3(data_list, filename):
    wb = Workbook()

    # Crear hoja 1 - Resumen de generaciones
    ws1 = wb.active
    ws1.title = "Resumen"

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Encabezado de la hoja 1
    ws1.cell(row=1, column=1, value="Generación").border = thin_border
    ws1.cell(row=1, column=2, value="Número de elementos").border = thin_border
    ws1.cell(row=1, column=3, value="Tiempo").border = thin_border
    ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 20

    # Llenado de la tabla en hoja 1
    for i in range(len(data_list)):
        idx = i + 2
        ws1.cell(row=idx, column=1, value=i + 1).border = thin_border
        ws1.cell(row=idx, column=2, value=len(data_list[i][0])).border = thin_border
        ws1.cell(row=idx, column=3, value=Utils.str_time(data_list[i][1])).border = thin_border
        ws1.cell(row=idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=3).alignment = Alignment(horizontal="center", vertical="center")

    # Crear hoja 2 - Datos
    ws2 = wb.create_sheet(title="Datos")

    col_offset = 0  # Para manejar el espaciado entre tablas

    for i in range(len(data_list)):
        col = 1 + col_offset  # Inicia la tabla en la siguiente posición

        # 1. Primera fila - Encabezado fusionado con bordes completos
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
        cell = ws2.cell(row=1, column=col, value=f"Generación {i + 1}")
        cell.alignment = Alignment(horizontal="center")
        apply_border(ws2, 1, col, 1, col + 4, thin_border)

        # 2. Segunda fila - "Tiempo" fusionado y número aleatorio con bordes
        ws2.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        cell = ws2.cell(row=2, column=col, value="Tiempo")
        cell.alignment = Alignment(horizontal="center")
        apply_border(ws2, 2, col, 2, col + 3, thin_border)
        ws2.cell(row=2, column=col + 4,
                 value=Utils.str_time(data_list[i][1])).border = thin_border  # Número aleatorio para tiempo

        # 3. Tercera fila - "Elementos" en la primera celda y número aleatorio con bordes
        ws2.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 3)
        cell = ws2.cell(row=3, column=col, value="Elementos")
        cell.alignment = Alignment(horizontal="center")
        apply_border(ws2, 3, col, 3, col + 3, thin_border)
        ws2.cell(row=3, column=col + 4,
                 value=len(data_list[i][0])).border = thin_border  # Número aleatorio para elementos

        # 4. Cuarta fila - Encabezado de columnas con bordes y ajuste de texto
        headers = ["Solapes", "Cohesion Teoria", "Equilibrio", "Cohesion Practicas", "Preferencias"]
        for j, header in enumerate(headers):
            cell = ws2.cell(row=4, column=col + j, value=header)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.column_dimensions[cell.column_letter].width = 15  # Ajuste de ancho automático

        # 5. Llenado de datos (valores aleatorios entre 0.0000 y 1.0000) con bordes
        for j in range(len(data_list[i][0])):
            row_idx = j + 5
            objetivos = [[int(parts[0])] + [float(x) for x in parts[1:]] for parts in
                         [str(data_list[i][0][j]).replace("(", "").replace(")", "").split()]]
            for col_idx, value in enumerate(objetivos[0]):
                cell = ws2.cell(row=row_idx, column=col + col_idx, value=round(value, 4))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 6. Aumentar el espacio para la siguiente tabla
        col_offset += 6  # 5 columnas + 1 de separación

    # Guardar archivo Excel
    wb.save(filename + "/" + "resumen.xlsx")
    print(f"Archivo guardado como {filename}")


def create_excel_resume(data_list, filename):
    wb = Workbook()

    # Crear hoja 1 - Resumen de generaciones
    ws1 = wb.active
    ws1.title = "Resumen"

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Encabezado de la hoja 1
    ws1.cell(row=1, column=1, value="Población").border = thin_border
    ws1.cell(row=1, column=2, value="Generaciones").border = thin_border
    ws1.cell(row=1, column=3, value="m_elemento").border = thin_border
    ws1.cell(row=1, column=4, value="m_teoria").border = thin_border
    ws1.cell(row=1, column=5, value="m_practica").border = thin_border
    ws1.cell(row=1, column=6, value="t_torneo").border = thin_border
    ws1.cell(row=1, column=7, value="particiones").border = thin_border
    ws1.cell(row=1, column=8, value="Número de elementos").border = thin_border
    ws1.cell(row=1, column=9, value="Tiempo").border = thin_border
    ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=7).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=1, column=9).alignment = Alignment(horizontal="center", vertical="center")

    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 25
    ws1.column_dimensions["D"].width = 25
    ws1.column_dimensions["E"].width = 25
    ws1.column_dimensions["F"].width = 25
    ws1.column_dimensions["G"].width = 25
    ws1.column_dimensions["H"].width = 25
    ws1.column_dimensions["I"].width = 25

    # Llenado de la tabla en hoja 1
    for i in range(len(data_list)):
        idx = i + 2

        ws1.cell(row=idx, column=1, value=data_list[i][0][0]).border = thin_border
        ws1.cell(row=idx, column=2, value=data_list[i][0][1]).border = thin_border
        ws1.cell(row=idx, column=3, value=data_list[i][0][2]).border = thin_border
        ws1.cell(row=idx, column=4, value=data_list[i][0][3]).border = thin_border
        ws1.cell(row=idx, column=5, value=data_list[i][0][4]).border = thin_border
        ws1.cell(row=idx, column=6, value=data_list[i][0][5]).border = thin_border
        ws1.cell(row=idx, column=7, value=data_list[i][0][6]).border = thin_border
        ws1.cell(row=idx, column=8, value=len(data_list[i][1])).border = thin_border
        ws1.cell(row=idx, column=9, value=Utils.str_time(data_list[i][2])).border = thin_border
        ws1.cell(row=idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=idx, column=9).alignment = Alignment(horizontal="center", vertical="center")

    # Crear hoja 2 - Datos
    ws2 = wb.create_sheet(title="Datos")

    col_offset = 0  # Para manejar el espaciado entre tablas

    for i in range(len(data_list)):
        col = 1 + col_offset  # Inicia la tabla en la siguiente posición

        # 1. Primera fila - Encabezado fusionado con bordes completos
        headers = ["poblacion", "generaciones", "m_elemento", "m_teoria", "m_practica", "t_torneo", "particiones"]
        header_values = [data_list[i][0][0], data_list[i][0][1], data_list[i][0][2], data_list[i][0][3],
                         data_list[i][0][4], data_list[i][0][5], data_list[i][0][6]]
        for j, header in enumerate(headers):
            cell = ws2.cell(row=1, column=col + j, value=header)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.column_dimensions[cell.column_letter].width = 20  # Ajuste de ancho automático

            cell = ws2.cell(row=2, column=col + j, value=header_values[j])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 3. Tercera fila - "Elementos" en la primera celda y número aleatorio con bordes
        ws2.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 3)
        cell = ws2.cell(row=3, column=col, value="Elementos")
        cell.alignment = Alignment(horizontal="center")
        apply_border(ws2, 3, col, 3, col + 3, thin_border)
        ws2.cell(row=3, column=col + 4,
                 value=len(data_list[i][1])).border = thin_border  # Número aleatorio para elementos

        # 4. Cuarta fila - Encabezado de columnas con bordes y ajuste de texto
        headers = ["Solapes", "Cohesion Teoria", "Equilibrio", "Cohesion Practicas", "Preferencias"]
        for j, header in enumerate(headers):
            cell = ws2.cell(row=4, column=col + j, value=header)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.column_dimensions[cell.column_letter].width = 15  # Ajuste de ancho automático

        # 5. Llenado de datos (valores aleatorios entre 0.0000 y 1.0000) con bordes
        for j in range(len(data_list[i][1])):
            row_idx = j + 5
            objetivos = [[int(parts[0])] + [float(x) for x in parts[1:]] for parts in
                         [str(data_list[i][1][j]).replace("(", "").replace(")", "").split()]]
            for col_idx, value in enumerate(objetivos[0]):
                cell = ws2.cell(row=row_idx, column=col + col_idx, value=round(value, 4))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 6. Aumentar el espacio para la siguiente tabla
        col_offset += 8  # 5 columnas + 1 de separación

    # Guardar archivo Excel
    wb.save(filename + ".xlsx")
    print(f"Archivo guardado como {filename}")



cruce = ["uniforme"]

poblacion = [100,300,500]
generaciones = [100,200,500]
mutacion_elemento = [0.00009]
mutacion_teoria = [0.0005,0.005]
mutacion_practica = [0.0005,0.005]
t_torneo = [2,5,10]
particiones = [6,7,8]


def iteracion(poblacion,generaciones,mutacion_elemento,mutacion_teoria,mutacion_practica,t_torneo, cruce, particion, t_total, resultados_totales, lock, carpeta, index):
    nsga3 = NSGA3(poblacion, generaciones, cruce, mutacion_elemento, mutacion_teoria, mutacion_practica, t_torneo, particion, carpeta, True)
    with lock:
        if not os.path.exists(carpeta):
            os.makedirs(carpeta)
    solucion = nsga3.ejecutar(lambda: False,True)
    create_excel_nsg3(solucion, carpeta)
    nsga3_p = [nsga3.tamaño_poblacion, nsga3.generaciones, nsga3.p_mutacion_alumno, nsga3.p_mutacion_teoria, nsga3.p_mutacion_practica, nsga3.tamaño_torneo, nsga3.particiones]
    solucion_l = []
    for s in solucion[-1][0]:
        solucion_l.append(str(s))
    time = 0
    for s in solucion:
        time += int(s[1])
    with lock:
        t_total.value+=nsga3.time
        resultados_totales[index] = [nsga3_p, solucion_l, time]



if __name__ == '__main__':
    n = len(poblacion) * len(generaciones) * len(mutacion_elemento) * len(mutacion_teoria) * len(mutacion_practica) * len(t_torneo) * len(particiones)
    carpeta = str(datetime.datetime.now())[:19].replace(':', '-') + '-nsga3'
    manager = mp.Manager()
    t_total = manager.Value('d', 0.0)
    resultados_totales = manager.list([None] * n)
    lock = manager.Lock()
    pool = mp.Pool(mp.cpu_count())
    async_results = []
    start_time = time.time()
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)
    index = 0
    for p in poblacion:
        for g in generaciones:
            for me in mutacion_elemento:
                for mt in mutacion_teoria:
                    for mp in mutacion_practica:
                        for t in t_torneo:
                            for c in cruce:
                                for pa in particiones:
                                    subcarpeta = f"{p}-{g}-{me}-{mt}-{mp}-{t}-{pa}"
                                    async_results.append(pool.apply_async(iteracion, args=(p, g, me, mt, mp, t, c, pa, t_total, resultados_totales, lock, carpeta+"/"+subcarpeta, index)))
                                    index+=1

    for result in async_results:
        result.get()

    pool.close()
    pool.join()

    end_time = time.time()
    create_excel_resume(resultados_totales, carpeta + "/resumen")
    total_time = end_time - start_time

    print(f"Tiempo Total NSGA-III: " + Utils.str_time(total_time))

