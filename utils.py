import random

import numpy as np

from alumno import alumno
from solucion import solucion
from openpyxl import Workbook
import pandas as pd


class Mutar:

    @staticmethod
    def mutar(alumno, p_mutacion_teoria, p_mutacion_practicas):
        horariosPath = "docs/horarios.xlsx"
        horarios = pd.read_excel(horariosPath)
        for matricula in alumno.matriculas_variables:
            if random.random() < p_mutacion_practicas:
                matricula.grupo_practicas = 1 if matricula.grupo_practicas == 2 else 2
                filtro2 = (horarios["CODIGO"] == matricula.cod_asignatura) & (
                        horarios["ID GRUPO"] == matricula.grupo) & (
                                  horarios["TEORÍA/PRÁCTICA"] == "P")
                matricula.horario_practicas = horarios.loc[filtro2].apply(lambda row: row['DÍA'] + '/' + row['HORARIO'],
                                                                          axis=1).tolist()
                matricula.horario_practicas.pop(abs(matricula.grupo_practicas - 2))

            if random.random() < p_mutacion_teoria:
                grupos = {10, 11, 12} if matricula.curso == 1 else {10, 11}
                matricula.grupo = random.choice(list(grupos.difference({matricula.grupo})))
                filtro = (horarios["CODIGO"] == matricula.cod_asignatura) & (
                            horarios["ID GRUPO"] == matricula.grupo) & (
                                 horarios["TEORÍA/PRÁCTICA"] == "T")
                matricula.horario_teoria = horarios.loc[filtro].apply(lambda row: row['DÍA'] + '/' + row['HORARIO'],
                                                                      axis=1).tolist()
                filtro2 = (horarios["CODIGO"] == matricula.cod_asignatura) & (
                        horarios["ID GRUPO"] == matricula.grupo) & (
                                  horarios["TEORÍA/PRÁCTICA"] == "P")
                matricula.horario_practicas = horarios.loc[filtro2].apply(lambda row: row['DÍA'] + '/' + row['HORARIO'],
                                                                          axis=1).tolist()
                matricula.horario_practicas.pop(abs(matricula.grupo_practicas - 2))
        return alumno


class Seleccion:

    @staticmethod
    def seleccion_fitness(poblacion,fitness):
        #probabilidades_seleccion = [fitness[i]/sum(fitness.values()) for i in range(len(poblacion))]
        probabilidades_seleccion = np.exp(list(fitness.values()))
        probabilidades_seleccion_n = probabilidades_seleccion / np.sum(probabilidades_seleccion)
        padre1 = poblacion[np.random.choice(len(poblacion),p=probabilidades_seleccion_n)]
        padre2 = poblacion[np.random.choice(len(poblacion),p=probabilidades_seleccion_n)]
        while padre1 == padre2:
            padre2 = poblacion[np.random.choice(len(poblacion),p=probabilidades_seleccion_n)]
        return padre1, padre2

    @staticmethod
    def seleccion_rango(poblacion,fitness):
        id_sorted = sorted(fitness, key=lambda x: fitness[x], reverse=True)
        probabilidades_seleccion = [(2*(len(poblacion)-(i+1)+1))/((len(poblacion)**2)+len(poblacion)) for i in range(len(poblacion))]
        padre1 = poblacion[np.random.choice(id_sorted, p=probabilidades_seleccion)]
        padre2 = poblacion[np.random.choice(id_sorted, p=probabilidades_seleccion)]
        while padre1 == padre2:
            padre2 = poblacion[np.random.choice(id_sorted, p=probabilidades_seleccion)]
        return padre1, padre2

    @staticmethod
    def seleccion_torneo(poblacion,fitness,k):

        torneo1 = random.sample([i for i in range(1, len(poblacion) + 1)], k)
        torneo2 = random.sample([i for i in range(1, len(poblacion) + 1)], k)
        id1 = sorted(torneo1, key=lambda x: fitness[x], reverse=True)[0]
        id2 = sorted(torneo2, key=lambda x: fitness[x], reverse=True)[0]
        while id1 == id2:
            torneo2 = random.sample([i for i in range(1, len(poblacion) + 1)], k)
            id2 = sorted(torneo2, key=lambda x: fitness[x], reverse=True)[0]
        padre1 = poblacion[id1]
        padre2 = poblacion[id2]
        return padre1, padre2

class Cruce:

    @staticmethod
    def cruce_punto(padres,listaAlumnos):
        nueva_generacion = []
        for pareja in padres:
            padre1, padre2 = pareja[0], pareja[1]
            division = int(len(padre1.alumnos) / 2)
            hijo1 = solucion(padre1.alumnos[:division]+padre2.alumnos[division:],listaAlumnos)
            hijo2 = solucion(padre2.alumnos[:division]+padre1.alumnos[division:],listaAlumnos)
            nueva_generacion.extend([hijo1,hijo2])
        return nueva_generacion

    @staticmethod
    def cruce_varios_puntos(padres,listaAlumnos):
        nueva_generacion = []
        for pareja in padres:
            padre1, padre2 = pareja[0], pareja[1]
            punto1 = int(len(padre1.alumnos) / 2) - 50
            punto2 = int(len(padre1.alumnos) / 2) + 50
            hijo1 = solucion(padre1.alumnos[:punto1] + padre2.alumnos[punto1:punto2] + padre1.alumnos[punto2:], listaAlumnos)
            hijo2 = solucion(padre2.alumnos[:punto1] + padre1.alumnos[punto1:punto2] + padre2.alumnos[punto2:], listaAlumnos)
            nueva_generacion.extend([hijo1, hijo2])
        return nueva_generacion

    @staticmethod
    def cruce_uniforme(padres,listaAlumnos):
        nueva_generacion = []
        mascara = [random.choice([False, True]) for _ in range(len(padres[0][0].alumnos))]
        for pareja in padres:
            padre1, padre2 = pareja[0], pareja[1]
            hijo1 = []
            hijo2 = []
            for i in range(len(padre1.alumnos)):
                hijo1.append(padre1.alumnos[i] if mascara[i] else padre2.alumnos[i])
                hijo2.append(padre2.alumnos[i] if mascara[i] else padre1.alumnos[i])

            nueva_generacion.extend([solucion(hijo1, listaAlumnos), solucion(hijo2, listaAlumnos)])

        return nueva_generacion


    @staticmethod
    def cruce_matriculas(padres,listaAlumnos):
        nueva_generacion = []
        for pareja in padres:
            padre1, padre2 = pareja[0], pareja[1]
            alumnos1 = []
            alumnos2 = []
            for i in range(len(padre1.alumnos)):
                if len(padre1.alumnos[i].matriculas_variables)>0:
                    b = (len(padre1.alumnos[i].matriculas_variables) - 1)
                    punto_corte = random.randint(0, b)
                    punto1 = int(len(padre1.alumnos[i].matriculas_variables) / 2) if punto_corte > int(len(padre1.alumnos[i].matriculas_variables) / 2) else punto_corte
                    punto2 = int(len(padre1.alumnos[i].matriculas_variables) / 2) if punto_corte <= int(len(padre1.alumnos[i].matriculas_variables) / 2) else punto_corte
                    matriculas_variables1 = padre1.alumnos[i].matriculas_variables[:punto1] + padre2.alumnos[i].matriculas_variables[punto1:punto2] + padre1.alumnos[i].matriculas_variables[punto2:]
                    matriculas_variables2 = padre2.alumnos[i].matriculas_variables[:punto1] + padre1.alumnos[i].matriculas_variables[punto1:punto2] + padre2.alumnos[i].matriculas_variables[punto2:]
                else:
                    matriculas_variables1 = []
                    matriculas_variables2 = []
                alumnos1.append(alumno(padre1.alumnos[i].nombre,matriculas_variables1,padre1.alumnos[i].matriculas_fijas))
                alumnos2.append(alumno(padre2.alumnos[i].nombre, matriculas_variables2, padre2.alumnos[i].matriculas_fijas))
            nueva_generacion.extend([solucion(alumnos1, listaAlumnos), solucion(alumnos2, listaAlumnos)])
        return nueva_generacion
class Sustitucion:

    @staticmethod
    def reemplazo(nueva_generacion):
        return nueva_generacion

    @staticmethod
    def elitismo(poblacionInicial,nueva_generacion,fitness):
        fitness_n = {i:nueva_generacion[i].calcular_fitness() for i in range(len(nueva_generacion))}
        max_id_old_g = max(fitness, key=lambda k: fitness[k])
        min_id_new_g = min(fitness_n, key=lambda k: fitness_n[k])
        nueva_generacion.pop(min_id_new_g)
        nueva_generacion.extend([poblacionInicial[max_id_old_g]])
        return nueva_generacion

    @staticmethod
    def truncamiento(poblacionInicial,nueva_generacion,fitness):
        fitness_n = {str(i) + 'b': nueva_generacion[i].calcular_fitness() for i in range(len(nueva_generacion))}
        fitness_n.update(fitness)
        n_individuos_id = sorted(fitness_n, key=lambda x: fitness_n[x], reverse=True)
        ng = []
        for i in range(len(poblacionInicial)):
            if not str(n_individuos_id[i])[-1].isalpha():
                individuo = poblacionInicial[n_individuos_id[i]]
                n_individuos_id.remove(str(n_individuos_id[i])+'b')
            else:
                individuo = nueva_generacion[int(str(n_individuos_id[i])[:-1])]
                n_individuos_id.remove(int(str(n_individuos_id[i])[:-1]))

            ng.append(individuo)
        return ng



class Utils:

    @staticmethod
    def print_solucion(poblacionInicial,fitness):
        print("fitness" + " ( " + "solapes" + ", " +
              "Cohesion teoria" + ", " +
              "Equilibrio grupos" + ", " +
              "Practicas pronto" + ", " +
              "Cohesion practicas" + ", " +
              "Preferencias" + " ) ")
        print("---------------------------------------------------------------------------------------------------------------------------------------")
        for e in range(len(poblacionInicial)):
            print(str(fitness[e]) + " ( " + str(poblacionInicial[e].solapes) + ", " +
                  str(poblacionInicial[e].tasa_cohesion) + ", " +
                  str(poblacionInicial[e].d_total) + ", " +
                  str(poblacionInicial[e].tasa_practicas_pronto) + ", " +
                  str(poblacionInicial[e].tasa_cohesion_practicas) + ", " +
                  str(poblacionInicial[e].tasa_preferencias) + " ) ")
        print("---------------------------------------------------------------------------------------------------------------------------------------")

    @staticmethod
    def export(listaAlumnos):
        wb = Workbook()
        ws = wb.active
        ws.append(["AÑO", "PLAN","CODIGO","ASIGNATURA","DNI","ALUMNO","GRUPO","GP"])
        for alumno in listaAlumnos:
            matriculas = alumno.matriculas_variables+alumno.matriculas_fijas
            for matricula in matriculas:
                ws.append(["2023-24","406",matricula.cod_asignatura,matricula.nombre_asignatura,"",alumno.nombre,matricula.grupo,matricula.grupo_practicas])
        wb.save('solucion.xlsx')


    @staticmethod
    def export_alumnos_clae(alumnosClaseInicio,alumnosClaseFinal,asignaturas):

        wb = Workbook()

        hoja1 = wb.active
        hoja1.title = "Configuracion Inicial"
        hoja2 = wb.create_sheet(title="Configuracion Final")

        hoja1.append(["NOMBRE","ALUMNOS","ESPERADOS POR GRUPO TEORIA", "ESPERADOS POR GRUPO PRACTICAS", "GRUPO 10", "GP1", "GP2", "GRUPO 11", "GP1", "GP2","GRUPO 12","GP1", "GP2",])
        hoja2.append(["NOMBRE","ALUMNOS","ESPERADOS POR GRUPO TEORIA", "ESPERADOS POR GRUPO PRACTICAS", "GRUPO 10", "GP1", "GP2", "GRUPO 11", "GP1", "GP2","GRUPO 12","GP1", "GP2",])

        alumnosClaseTeoriaInicio,alumnosClasePracticasInicio = alumnosClaseInicio[0], alumnosClaseInicio[1]
        alumnosClaseTeoriaFinal, alumnosClasePracticasFinal = alumnosClaseFinal[0], alumnosClaseFinal[1]

        for clave in alumnosClaseTeoriaInicio.keys():

            cod_asignatura = clave[0]
            filtro = (asignaturas["COD. ASIG"] == cod_asignatura)
            datos_asignatura = asignaturas.loc[filtro,["NOMBRE ASIGNATURA", "CURSO"]]
            if (datos_asignatura["CURSO"] == 1).any():
                hoja1.append([datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaInicio[clave].values()), int(sum(alumnosClaseTeoriaInicio[clave].values())/3),int((sum(alumnosClaseTeoriaInicio[clave].values())/3)/2),alumnosClaseTeoriaInicio[clave][10],alumnosClasePracticasInicio[clave[0]][10][1],alumnosClasePracticasInicio[clave[0]][10][2],
                              alumnosClaseTeoriaInicio[clave][11],alumnosClasePracticasInicio[clave[0]][11][1],alumnosClasePracticasInicio[clave[0]][11][2],alumnosClaseTeoriaInicio[clave][12],alumnosClasePracticasInicio[clave[0]][12][1],alumnosClasePracticasInicio[clave[0]][12][2]])
                hoja2.append([datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaFinal[clave].values()),
                              int(sum(alumnosClaseTeoriaFinal[clave].values()) / 3),
                              int((sum(alumnosClaseTeoriaFinal[clave].values()) / 3) / 2),
                              alumnosClaseTeoriaFinal[clave][10], alumnosClasePracticasFinal[clave[0]][10][1],
                              alumnosClasePracticasFinal[clave[0]][10][2],
                              alumnosClaseTeoriaFinal[clave][11], alumnosClasePracticasFinal[clave[0]][11][1],
                              alumnosClasePracticasFinal[clave[0]][11][2], alumnosClaseTeoriaFinal[clave][12],
                              alumnosClasePracticasFinal[clave[0]][12][1],
                              alumnosClasePracticasFinal[clave[0]][12][2]])
            else:
                hoja1.append([datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaInicio[clave].values()),
                              int(sum(alumnosClaseTeoriaInicio[clave].values()) / 2),
                              int((sum(alumnosClaseTeoriaInicio[clave].values()) / 2) / 2),
                              alumnosClaseTeoriaInicio[clave][10], alumnosClasePracticasInicio[clave[0]][10][1],
                              alumnosClasePracticasInicio[clave[0]][10][2],
                              alumnosClaseTeoriaInicio[clave][11], alumnosClasePracticasInicio[clave[0]][11][1],
                              alumnosClasePracticasInicio[clave[0]][11][2], "",
                              "",
                              ""])
                hoja2.append([datos_asignatura.iloc[0]["NOMBRE ASIGNATURA"], sum(alumnosClaseTeoriaFinal[clave].values()),
                              int(sum(alumnosClaseTeoriaFinal[clave].values()) / 2),
                              int((sum(alumnosClaseTeoriaFinal[clave].values()) / 2) / 2),
                              alumnosClaseTeoriaFinal[clave][10], alumnosClasePracticasFinal[clave[0]][10][1],
                              alumnosClasePracticasFinal[clave[0]][10][2],
                              alumnosClaseTeoriaFinal[clave][11], alumnosClasePracticasFinal[clave[0]][11][1],
                              alumnosClasePracticasFinal[clave[0]][11][2], "",
                              "",
                              ""])

        wb.save("estudiantes_asignatura.xlsx")